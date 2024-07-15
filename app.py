import sys

import os
from openpyxl import load_workbook
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTableWidget, QMainWindow, \
    QHBoxLayout, QFileDialog, QLabel, QTableWidgetItem, QDialog, QHeaderView, QScrollArea, QInputDialog, QMessageBox
from PyQt5.QtChart import QChartView, QChart, QScatterSeries, QValueAxis
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QIcon
from launch_model import *
from architecture_of_model.arch_model_2 import SimpleNet
import pandas as pd
import torch

class FaqWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Часто задаваемые вопросы")
        faq_icon = QIcon("assets/faq_icon.png")
        self.setWindowIcon(faq_icon)
        self.setFixedSize(900, 250)
        text = open('assets/faq_text.txt', 'r', encoding="UTF-8")

        label = QLabel(text.read(), self)
        label.setWordWrap(True)
        label_font = QFont("Arial", 12)
        label.setFont(label_font)
        layout = QVBoxLayout()
        layout.addWidget(label)
        self.setLayout(layout)

class DataInputWindow(QDialog):
    def __init__(self, num_rows):
        super().__init__()
        self.setWindowTitle("Ввод данных")
        self.resize(800, 500)

        self.button_load_data = QPushButton("Загрузить данные", self)
        self.button_load_data.setFixedSize(150, 30)
        self.button_load_data.move(20, 457)
        self.button_load_data.clicked.connect(self.save_data_to_array)

        # Создание таблицы
        self.manual_table = QTableWidget()
        self.manual_table.setColumnCount(10)  # 3 столбца
        self.manual_table.setRowCount(num_rows)
        self.manual_table.setHorizontalHeaderLabels(["age", "gender", "LF/HF", "%VLF", "%LF", "%HF", "LF/HF", "%VLF", "%LF", "%HF"])
        self.manual_table.horizontalHeader().setDefaultSectionSize(75)

        self.manual_table.itemChanged.connect(self.validate_column)

        # Размещение элементов
        layout = QVBoxLayout()
        layout.addWidget(self.manual_table)
        layout.addSpacing(40)
        self.setLayout(layout)

    def validate_column(self, item):
        column = item.column()
        if column == 1:
            text = item.text()
            if text in ("М", "Ж", "м", "ж"):
                item.setText(text)
            else:
                QMessageBox.warning(self, "Некорректные данные", "Введите 'М' или 'Ж' в этот столбец.")
                item.setText("М")

        if column in [0, 2, 3, 4, 5, 6, 7, 8, 9]:
            try:
                float(item.text())
            except ValueError:
                # Выводим предупреждение
                QMessageBox.warning(self, "Некорректные данные", "Введите число в этот столбец.")
                item.setText("0")
    def save_data_to_array(self, filename):
        # Получаем количество строк и столбцов в таблице
        num_rows = self.manual_table.rowCount()
        num_cols = self.manual_table.columnCount()

        # Создаем массив для хранения данных
        manual_input = []

        # Заполняем массив данными из таблицы
        for row in range(num_rows):
            row_data = []
            for col in range(num_cols):
                item = self.manual_table.item(row, col)
                if item:
                    row_data.append(item.text())
                else:
                    QMessageBox.warning(self, "Некорректные данные", "Заполните все ячейки")
                    return
            manual_input.append(row_data)

        # Теперь у вас есть массив data_array, содержащий данные из таблицы
        print("Данные в массиве:", manual_input)

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Заполнение данных в таблицу Excel
        worksheet.append(["age", "gender", "LF/HF", "%VLF", "%LF", "%HF", "LF/HF", "%VLF", "%LF", "%HF"])
        for row in manual_input:
            worksheet.append(row)

        # Сохранение файла
        workbook.save("manual_data.xlsx")

        # Перемещение файла с заменой существующего
        source_path = "manual_data.xlsx"
        destination_path = os.path.join("assets", "manual_data.xlsx")

        try:
            if os.path.exists(destination_path):
                os.remove(destination_path)
                print(f"Файл {destination_path} удален, так как уже существует.")
            os.rename(source_path, destination_path)
            print("Файл успешно перемещен в папку assets")
        except FileNotFoundError:
            print("Файл не найден. Проверьте имя файла и путь.")
        except OSError as e:
            print(f"Ошибка перемещения файла: {e}")

        QMessageBox.information(self, "Успешно", "Данные успешно сохранены в файл 'manual_data.xlsx'!")

class ChartWidget(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle('Rocket')
        self.setGeometry(100, 100, 600, 400)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QHBoxLayout()
        central_widget.setLayout(layout)

        # Кнопки
        close_button = QPushButton('✖️', self)
        close_button.setFixedSize(30, 30)
        close_button.clicked.connect(self.close)
        close_button.move(1885, 5)

        faq_button = QPushButton('FAQ', self)
        faq_button.setFixedSize(90, 30)
        faq_button.move(1765, 5)
        faq_button.clicked.connect(self.open_faq)

        minimize_button = QPushButton('‒', self)
        minimize_button.setFixedSize(30, 30)
        minimize_button.clicked.connect(self.showMinimized)
        minimize_button.move(1855, 5)

        upload_button = QPushButton('Загрузить файл', self)
        upload_button.setFixedSize(220, 63)
        upload_button.move(23, 1000)
        upload_button.clicked.connect(self.get_filename)

        button_manual_input = QPushButton('Ввести данные вручную', self)
        button_manual_input.setFixedSize(220, 63)
        button_manual_input.move(257, 1000)
        #upload_table_button.clicked.connect(self.make_predict)
        # Кнопка "Ввести данные вручную"
        button_manual_input.clicked.connect(self.open_data_input_window)

        start_button = QPushButton('Спрогнозировать', self)
        start_button.setFixedSize(220, 63)
        start_button.move(490, 1000)
        start_button.clicked.connect(self.chart_points)

        self.label = QLabel("Результат реабилитации на основе табличных данных", self)
        self.label.move(30, 900)
        self.label.setFixedSize(675, 70)
        font = QFont("Arial", 12)
        self.label.setFont(font)

        # Создаем график
        self.chart = QChart()
        #self.chart.chartType()
        self.chart.setTitle("Качество восстановления от показателей волн")

        # Создание таблицы
        self.table = QTableWidget(1000, 11)
        self.table.setFixedSize(685, 825)
        #self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["age", "gender", "LF/HF", "%VLF", "%LF", "%HF", "LF/HF", "%VLF", "%LF", "%HF", "Итог"])
        table_columns = self.table.horizontalHeader()
        table_columns.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table_columns.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        table_columns.setSectionResizeMode(10, QHeaderView.Stretch)
        self.table.horizontalHeader().setDefaultSectionSize(57)
        #layout.addWidget(self.table)

        # # Создание виджета для графика
        self.chart_view = QChartView(self.chart)
        self.chart_view.setFixedSize(1200, 1000)
        layout.addSpacing(15)
        layout.addWidget(self.chart_view, 0, Qt.AlignRight)

        # Настройка шрифта
        font = QFont()
        font.setPointSize(12)
        self.chart.setFont(font)

        layout_top = QVBoxLayout()
        layout_top.addSpacing(40)
        layout_top.addWidget(self.table, 0, Qt.AlignTop)

        # Добавляем таблицу в макет
        layout = self.centralWidget().layout()
        layout.addLayout(layout_top)
        layout.addWidget(self.chart_view)

        self.show()

    def open_faq(self):
        faq_window = FaqWindow()
        faq_window.exec_()

    def open_data_input_window(self):
        DataInputWindow().exec_()

    def open_data_input_window(self):
        # Получение количества строк от пользователя
        num_rows, ok = QInputDialog.getInt(self, "Ввод данных", "Введите количество строк:", min=1)

        if ok:
            DataInputWindow(num_rows).exec_()

    def get_filename(self):
        filename, filetype = QFileDialog.getOpenFileName(
            self,
            "Открыть файл",
            ".",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if filename:
            self.read_excel(filename)
        return filename

    def make_predict(self, filename):
        lm = LaunchModel() # определяем экземпляр класса
        data_lm = lm.processing_data_file(filename) # обработка данных и получения тензора
        print(data_lm)
        predictions = lm.launch_model_2(X=data_lm)
        print(predictions)
        return predictions

    def read_excel(self, filename: str):
        data = load_workbook(filename)
        worksheet = data.active

        if worksheet.cell(row = 2, column = 1).value == None:
            QMessageBox.information(self, "Ошибка", "В файле неподходящие входные данные!")
        else:
            prediction = self.make_predict(filename)
            prediction = [[int(i) for i in inner_arr] for inner_arr in prediction]

            # Определение количества строк и столбцов в таблице
            row = worksheet.max_row
            cols = worksheet.max_column

            rows = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        rows += 1
                        break

            # Настройка таблицы PyQt5
            #self.table = QTableWidget(self)
            self.table.setRowCount(rows-1)
            self.table.setColumnCount(11)
            self.table.setHorizontalHeaderLabels(list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0])

            self.table.setFixedSize(685, 825)

            # Заполнение таблицы данными из Excel
            for row_index in range(1, rows+1):
                for col_index in range(cols):
                    cell_value = worksheet.cell(row=row_index + 1, column=col_index + 1).value
                    item = QTableWidgetItem(str(cell_value))
                    self.table.setItem(row_index-1, col_index, item)

            # worksheet.column_dimensions["J"].custom_width = 20  # Изменяем ширину столбца
            # worksheet.cell(row=1, column=10).value = "Итог"  # Изменяем название столбца
            for row_index in range(1, rows+1):
                if row_index <= len(prediction):
                    item = QTableWidgetItem(str(*prediction[row_index-1]))
                    self.table.setItem(row_index-1, 10, item)

            count_1 = prediction.count([1])
            count_6 = 0
            count_7 = 0
            total_6 = 0
            total_7 = 0
            for row in range(self.table.rowCount()):
                if int(self.table.item(row, 0).text()) < 7:
                    total_6 += 1
                    if int(self.table.item(row, 10).text()) == 1:
                        count_6 += 1
                else:
                    total_7 += 1
                    if int(self.table.item(row, 10).text()) == 1:
                        count_7 += 1
            text = f"Общий прогноз успешной реабилитации: {round(count_1 / len(prediction), 4) * 100}% ({count_1}/{len(prediction)})\n"
            if total_6 != 0:
                text += f"Пациентов младше 7-ми лет: {round(count_6 / total_6, 4) * 100}% ({count_6}/{total_6})\n"
            if total_7 != 0:
                text += f"Пациентов 7-ми лет и старше: {round(count_7 / total_7, 4) * 100}% ({count_7}/{total_7})"
            self.label.setText(text)

    def chart_points(self):
        # Создание точек на графике
        if self.table.item(0,0) == None:
            QMessageBox.information(self, "Ошибка", "Вы не выбрали файл!")
        else:
            self.chart.removeAllSeries()
            series_6 = QScatterSeries()
            series_6.setName("Младше 7 лет")
            series_6.setColor(Qt.GlobalColor.red)
            series_6.setBorderColor(Qt.GlobalColor.black)
            series_6.setMarkerSize(10)

            series_7 = QScatterSeries()
            series_7.setName("Старше 7 лет")
            series_7.setColor(Qt.GlobalColor.blue)
            series_7.setBorderColor(Qt.GlobalColor.black)
            series_7.setMarkerSize(10)

            k = 1
            for i in range(self.table.rowCount()):
                x = k
                y = float(self.table.item(i, 10).text())
                if (float(self.table.item(i, 0).text()) <= 6):
                    series_6.append(x, y)
                else:
                    series_7.append(x, y)
                k += 1
            self.chart.addSeries(series_6)
            self.chart.addSeries(series_7)

            # Настройка осей
            self.chart.createDefaultAxes()
            self.chart.axisX().setTitleText("Индекс пациента")
            self.chart.axisY().setTitleText("Предсказание")
            #self.chart.axisY().setTitleText("Предсказание (0 - плохой прогноз, 1 - хороший прогноз)")

            self.chart.axisY().setRange(-0.1, 1.1)
            self.chart.axisY().setLabelFormat("%0.0f")
            self.chart.axisY().setTickType(QValueAxis.TickType.TicksDynamic)
            self.chart.axisY().setTickInterval(1)

            self.chart.axisX().setRange(0, self.table.rowCount() + 1)
            self.chart.axisX().setLabelFormat("%0.0f")
            self.chart.axisX().setTickType(QValueAxis.TickType.TicksDynamic)
            self.chart.axisX().setTickInterval(2)


def main():
    app = QApplication(sys.argv)
    main_window = ChartWidget()
    main_window.showFullScreen()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
